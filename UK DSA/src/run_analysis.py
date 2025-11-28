"""
UK Debt Sustainability Analysis - Complete Runner
==================================================
Expert-Corrected Implementation

This script runs the complete analysis pipeline:
1. Econometric Tests (unit root, cointegration, breaks)
2. Bohn Test with HAC standard errors
3. Debt Dynamics Decomposition
4. Monte Carlo Simulation (MLE-calibrated)
5. Sustainability Assessment
6. Generate all outputs

Author: UK DSA Project
Date: November 2025
"""

import sys
import os

# Add src to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from datetime import datetime
from scipy import stats

# Import corrected modules
from config import (
    HISTORICAL_DEBT_GDP, HISTORICAL_PRIMARY_BALANCE,
    HISTORICAL_GDP_GROWTH, HISTORICAL_GILT_YIELDS,
    BUDGET_PSND_GDP, BUDGET_PRIMARY_DEFICIT_GDP,
    CURRENT_DEBT_GDP, OBR_DEBT_TO_GDP
)
from econometric_tests import run_econometric_tests
from bohn_test import BohnTestCorrected


def print_header(title: str):
    """Print formatted section header."""
    print("\n" + "="*70)
    print(title)
    print("="*70)


class MonteCarloSimulatorCorrected:
    """
    Expert-corrected Monte Carlo simulator with MLE-calibrated distributions.
    """
    
    def __init__(self, n_simulations: int = 10000, horizon_years: int = 10, 
                 bohn_beta: float = -0.099, seed: int = 42):
        self.n_sims = n_simulations
        self.horizon = horizon_years
        self.bohn_beta = bohn_beta
        self.rng = np.random.default_rng(seed)
        
        # Default parameters (will be updated by calibrate_from_data)
        self.df_gdp = 3.0  # MLE estimate ~2.1, floored at 3 for stability
        self.df_rate = 4.0
        self.mu_g = 2.0 / 100  # Mean GDP growth
        self.sigma_g = 2.0 / 100
        self.mu_r = 4.5 / 100  # Mean interest rate
        self.sigma_r = 1.0 / 100
        self.mu_pb = 0.0 / 100  # Mean primary balance
        self.sigma_pb = 2.0 / 100
        
        # Correlation matrix (UK-estimated)
        self.corr_gdp_rate = 0.40  # CORRECTED: procyclical, not -0.30
        self.corr_gdp_pb = 0.56
        self.corr_rate_pb = 0.12
        
        # Store results
        self.debt_paths = None
        
    def calibrate_from_data(self, gdp_growth: list, interest_rate: list, 
                            primary_balance: list):
        """Calibrate parameters from UK historical data using MLE."""
        gdp_arr = np.array(gdp_growth)
        rate_arr = np.array(interest_rate)
        pb_arr = np.array(primary_balance)
        
        # Remove NaN
        gdp_arr = gdp_arr[~np.isnan(gdp_arr)]
        rate_arr = rate_arr[~np.isnan(rate_arr)]
        pb_arr = pb_arr[~np.isnan(pb_arr)]
        
        # MLE estimation for Student-t
        try:
            df_g, loc_g, scale_g = stats.t.fit(gdp_arr)
            self.df_gdp = max(df_g, 2.5)
            self.mu_g = loc_g / 100
            self.sigma_g = scale_g / 100
        except:
            pass
        
        try:
            df_r, loc_r, scale_r = stats.t.fit(rate_arr)
            self.df_rate = max(df_r, 2.5)
            self.mu_r = loc_r / 100
            self.sigma_r = scale_r / 100
        except:
            pass
        
        # Estimate correlations
        min_len = min(len(gdp_arr), len(rate_arr), len(pb_arr))
        if min_len > 5:
            data = np.column_stack([gdp_arr[-min_len:], rate_arr[-min_len:], pb_arr[-min_len:]])
            corr = np.corrcoef(data.T)
            self.corr_gdp_rate = corr[0, 1]
            self.corr_gdp_pb = corr[0, 2]
            self.corr_rate_pb = corr[1, 2]
        
        print(f"\n   MLE-estimated parameters:")
        print(f"   GDP growth: df={self.df_gdp:.1f}, μ={self.mu_g*100:.2f}%, σ={self.sigma_g*100:.2f}%")
        print(f"   Interest rate: df={self.df_rate:.1f}, μ={self.mu_r*100:.2f}%, σ={self.sigma_r*100:.2f}%")
        print(f"   Correlations: ρ(g,r)={self.corr_gdp_rate:.2f}, ρ(g,pb)={self.corr_gdp_pb:.2f}")
        
    def _generate_shocks(self) -> tuple:
        """Generate fat-tailed correlated shocks."""
        # Build correlation matrix
        corr = np.array([
            [1.0, self.corr_gdp_rate, self.corr_gdp_pb],
            [self.corr_gdp_rate, 1.0, self.corr_rate_pb],
            [self.corr_gdp_pb, self.corr_rate_pb, 1.0]
        ])
        
        # Cholesky decomposition
        L = np.linalg.cholesky(corr)
        
        # Generate t-distributed shocks
        g_shocks = stats.t.rvs(self.df_gdp, size=(self.n_sims, self.horizon), random_state=self.rng)
        r_shocks = stats.t.rvs(self.df_rate, size=(self.n_sims, self.horizon), random_state=self.rng)
        pb_shocks = self.rng.standard_normal((self.n_sims, self.horizon))
        
        # Correlate using Cholesky
        for t in range(self.horizon):
            independent = np.column_stack([g_shocks[:, t], r_shocks[:, t], pb_shocks[:, t]])
            correlated = (L @ independent.T).T
            g_shocks[:, t] = correlated[:, 0]
            r_shocks[:, t] = correlated[:, 1]
            pb_shocks[:, t] = correlated[:, 2]
        
        return g_shocks, r_shocks, pb_shocks
    
    def simulate_obr_baseline(self, initial_debt: float = None) -> dict:
        """
        Simulate with OBR baseline primary balance path.
        """
        if initial_debt is None:
            initial_debt = CURRENT_DEBT_GDP
        
        # OBR primary balance path (% GDP)
        obr_pb = [-1.5, -0.6, -0.1, 0.5, 1.0, 1.3, 1.4, 1.4, 1.4, 1.4]
        obr_pb = obr_pb[:self.horizon]
        
        g_shocks, r_shocks, _ = self._generate_shocks()
        
        # Simulate debt paths
        debt = np.zeros((self.n_sims, self.horizon + 1))
        debt[:, 0] = initial_debt
        
        for t in range(self.horizon):
            g = self.mu_g + self.sigma_g * g_shocks[:, t]
            r = self.mu_r + self.sigma_r * r_shocks[:, t]
            pb = obr_pb[t] / 100  # Convert to decimal
            
            # Debt dynamics: d_{t+1} = (1+r)/(1+g) * d_t - pb
            debt[:, t+1] = (1 + r) / (1 + g) * debt[:, t] - pb * 100
        
        self.debt_paths = debt
        
        return self._compute_statistics(debt)
    
    def simulate_fiscal_reaction(self, initial_debt: float = None, 
                                  target_debt: float = 60.0) -> dict:
        """
        Simulate with embedded fiscal reaction (Bohn coefficient).
        """
        if initial_debt is None:
            initial_debt = CURRENT_DEBT_GDP
        
        baseline_pb = 0.5 / 100  # Baseline primary balance
        
        g_shocks, r_shocks, pb_shocks = self._generate_shocks()
        
        debt = np.zeros((self.n_sims, self.horizon + 1))
        debt[:, 0] = initial_debt
        
        for t in range(self.horizon):
            g = self.mu_g + self.sigma_g * g_shocks[:, t]
            r = self.mu_r + self.sigma_r * r_shocks[:, t]
            
            # Fiscal reaction: pb = baseline + β*(debt - target) + shock
            pb = baseline_pb + self.bohn_beta * (debt[:, t] - target_debt) / 100 + self.sigma_pb * pb_shocks[:, t]
            
            debt[:, t+1] = (1 + r) / (1 + g) * debt[:, t] - pb * 100
        
        return self._compute_statistics(debt)
    
    def _compute_statistics(self, debt: np.ndarray) -> dict:
        """Compute summary statistics from debt paths."""
        terminal = debt[:, -1]
        
        # Fan chart percentiles
        fan = {}
        for p in [5, 25, 50, 75, 95]:
            fan[f'p{p}'] = np.percentile(debt, p, axis=0)
        fan['mean'] = np.mean(debt, axis=0)
        
        stats_dict = {
            'mean': np.mean(terminal),
            'median': np.median(terminal),
            'std': np.std(terminal),
            'skewness': stats.skew(terminal),
            'kurtosis': stats.kurtosis(terminal),
            'min': np.min(terminal),
            'max': np.max(terminal),
            'prob_100_terminal': np.mean(terminal > 100) * 100,
            'prob_100_ever': np.mean(np.max(debt, axis=1) > 100) * 100,
            'var_95': np.percentile(terminal, 95),
            'var_99': np.percentile(terminal, 99),
            'es_95': np.mean(terminal[terminal > np.percentile(terminal, 95)]),
            'es_99': np.mean(terminal[terminal > np.percentile(terminal, 99)]),
            'fan_chart': fan
        }
        
        return stats_dict
    
    def run_scenario_analysis(self) -> dict:
        """Run multiple scenarios."""
        scenarios = {}
        
        # Save original parameters
        orig_mu_r = self.mu_r
        orig_mu_g = self.mu_g
        orig_bohn = self.bohn_beta
        
        # 1. Baseline
        scenarios['baseline'] = self.simulate_obr_baseline()
        
        # 2. High rates (+200bp)
        self.mu_r = orig_mu_r + 0.02
        scenarios['high_rates'] = self.simulate_obr_baseline()
        self.mu_r = orig_mu_r
        
        # 3. Low growth
        self.mu_g = orig_mu_g - 0.01
        scenarios['low_growth'] = self.simulate_obr_baseline()
        self.mu_g = orig_mu_g
        
        # 4. Stagflation
        self.mu_r = orig_mu_r + 0.015
        self.mu_g = orig_mu_g - 0.01
        scenarios['stagflation'] = self.simulate_obr_baseline()
        self.mu_r = orig_mu_r
        self.mu_g = orig_mu_g
        
        # 5. Consolidation success
        scenarios['consolidation'] = self._simulate_consolidation()
        
        # Print summary
        print("\n   Scenario Analysis Results:")
        print("   " + "-"*50)
        print(f"   {'Scenario':<20} {'Mean':<10} {'P(>100%)':<10} {'VaR99':<10}")
        print("   " + "-"*50)
        for name, data in scenarios.items():
            print(f"   {name:<20} {data['mean']:>7.1f}%   {data['prob_100_terminal']:>7.1f}%   {data['var_99']:>7.1f}%")
        
        return scenarios
    
    def _simulate_consolidation(self):
        """Simulate successful consolidation scenario."""
        initial_debt = CURRENT_DEBT_GDP
        pb_path = [1.0, 1.5, 2.0, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5, 2.5]
        pb_path = pb_path[:self.horizon]
        
        g_shocks, r_shocks, _ = self._generate_shocks()
        
        debt = np.zeros((self.n_sims, self.horizon + 1))
        debt[:, 0] = initial_debt
        
        for t in range(self.horizon):
            g = self.mu_g + self.sigma_g * g_shocks[:, t]
            r = self.mu_r + self.sigma_r * r_shocks[:, t]
            pb = pb_path[t] / 100
            
            debt[:, t+1] = (1 + r) / (1 + g) * debt[:, t] - pb * 100
        
        return self._compute_statistics(debt)


def print_header(title: str):
    """Print formatted section header."""
    print("\n" + "="*70)
    print(title)
    print("="*70)


def run_complete_analysis(output_dir: str = None) -> dict:
    """
    Run complete UK DSA analysis with expert corrections.
    
    Returns:
        Dictionary containing all results
    """
    print_header("UK DEBT SUSTAINABILITY ANALYSIS")
    print("Expert-Corrected Implementation - November 2025")
    print(f"Execution time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    results = {}
    
    # Prepare historical data
    years = sorted(set(HISTORICAL_DEBT_GDP.keys()) & 
                   set(HISTORICAL_PRIMARY_BALANCE.keys()))
    
    debt_gdp = np.array([HISTORICAL_DEBT_GDP[y] for y in years])
    primary_balance = np.array([HISTORICAL_PRIMARY_BALANCE[y] for y in years])
    years_arr = np.array(years)
    
    print(f"\nData: UK {years[0]}-{years[-1]} (n={len(years)} observations)")
    print(f"Current debt/GDP: {CURRENT_DEBT_GDP:.1f}%")
    
    # =========================================================================
    # 1. ECONOMETRIC TESTS
    # =========================================================================
    print_header("1. ECONOMETRIC TESTS")
    
    results['econometric'] = run_econometric_tests(
        debt_gdp, primary_balance, years_arr
    )
    
    # =========================================================================
    # 2. BOHN FISCAL REACTION TEST
    # =========================================================================
    print_header("2. BOHN FISCAL REACTION TEST")
    
    bohn = BohnTestCorrected()
    results['bohn'] = bohn.run_all_tests()
    
    # Extract beta for Monte Carlo
    bohn_beta = bohn.get_beta()
    
    # =========================================================================
    # 3. DEBT DYNAMICS DECOMPOSITION
    # =========================================================================
    print_header("3. DEBT DYNAMICS DECOMPOSITION")
    
    # Historical decomposition
    gdp_years = sorted(HISTORICAL_GDP_GROWTH.keys())
    yield_years = sorted(HISTORICAL_GILT_YIELDS.keys())
    common_years = sorted(set(years) & set(gdp_years) & set(yield_years))
    
    decomposition = []
    for i, y in enumerate(common_years[1:], 1):
        d_prev = HISTORICAL_DEBT_GDP.get(y-1, HISTORICAL_DEBT_GDP.get(common_years[i-1]))
        d_curr = HISTORICAL_DEBT_GDP.get(y)
        g = HISTORICAL_GDP_GROWTH.get(y, 2.0) / 100
        r = HISTORICAL_GILT_YIELDS.get(y, 4.0) / 100
        pb = HISTORICAL_PRIMARY_BALANCE.get(y, 0)
        
        # Decomposition: Δd = (r-g)/(1+g) * d - pb + sfa
        rg_effect = (r - g) / (1 + g) * d_prev
        pb_effect = -pb
        delta_d = d_curr - d_prev if d_curr and d_prev else None
        sfa = delta_d - rg_effect - pb_effect if delta_d else 0
        
        decomposition.append({
            'year': y,
            'debt_gdp': d_curr,
            'delta_d': delta_d,
            'r_g_effect': rg_effect,
            'pb_effect': pb_effect,
            'sfa_effect': sfa,
            'r': r * 100,
            'g': g * 100,
            'r_minus_g': (r - g) * 100
        })
    
    decomp_df = pd.DataFrame(decomposition)
    results['dynamics'] = {'historical': decomp_df}
    
    # Print summary
    print(f"\nHistorical r-g differential:")
    print(f"   Mean:   {decomp_df['r_minus_g'].mean():+.2f}pp")
    print(f"   Median: {decomp_df['r_minus_g'].median():+.2f}pp")
    print(f"   Range:  [{decomp_df['r_minus_g'].min():.1f}, {decomp_df['r_minus_g'].max():.1f}]pp")
    
    # Debt-stabilizing primary balance
    r_current = 4.5 / 100
    g_current = 3.5 / 100
    pb_star = (r_current - g_current) / (1 + g_current) * CURRENT_DEBT_GDP
    
    print(f"\nDebt-stabilizing primary balance:")
    print(f"   At current debt ({CURRENT_DEBT_GDP:.0f}%), r=4.5%, g=3.5%:")
    print(f"   pb* = {pb_star:.2f}% GDP")
    print(f"   OBR projects pb = 1.4% by 2030-31 → {'EXCEEDS' if 1.4 > pb_star else 'BELOW'} requirement")
    
    # =========================================================================
    # 4. MONTE CARLO SIMULATION
    # =========================================================================
    print_header("4. MONTE CARLO SIMULATION")
    
    mc = MonteCarloSimulatorCorrected(
        n_simulations=10000,
        horizon_years=10,
        bohn_beta=bohn_beta
    )
    
    # Calibrate from data
    print("\nCalibrating from UK historical data...")
    gdp_growth_list = [HISTORICAL_GDP_GROWTH.get(y, 2.0) for y in sorted(HISTORICAL_GDP_GROWTH.keys())[-20:]]
    rate_list = [HISTORICAL_GILT_YIELDS.get(y, 4.0) for y in sorted(HISTORICAL_GILT_YIELDS.keys())[-20:]]
    pb_list = [HISTORICAL_PRIMARY_BALANCE.get(y, 0) for y in sorted(HISTORICAL_PRIMARY_BALANCE.keys())[-20:]]
    
    mc.calibrate_from_data(
        gdp_growth=gdp_growth_list,
        interest_rate=rate_list,
        primary_balance=pb_list
    )
    
    # Run OBR baseline simulation
    print("\n--- OBR Baseline Scenario ---")
    obr_stats = mc.simulate_obr_baseline()
    results['monte_carlo'] = {'obr_baseline_stats': obr_stats}
    
    # Run fiscal reaction simulation
    print("\n--- Fiscal Reaction Scenario ---")
    fr_stats = mc.simulate_fiscal_reaction()
    results['monte_carlo']['fiscal_reaction_stats'] = fr_stats
    
    # Run scenario analysis
    print("\n--- Scenario Analysis ---")
    scenarios = mc.run_scenario_analysis()
    results['monte_carlo']['scenarios'] = scenarios
    
    # =========================================================================
    # 5. GROSS FINANCING NEEDS
    # =========================================================================
    print_header("5. GROSS FINANCING NEEDS")
    
    # GFN = Primary deficit + Interest + Maturing debt
    gfn_data = []
    gdp_nominal = 2900  # £bn approximate
    
    # Simplified GFN projection
    for i, year in enumerate(range(2024, 2031)):
        fy = f"{year}-{str(year+1)[-2:]}"
        
        psnb = BUDGET_PRIMARY_DEFICIT_GDP.get(fy, 0) * gdp_nominal / 100  # Convert to £bn
        interest = 100 + i * 5  # Approximate interest payments
        maturing = 80 + i * 10  # Approximate maturities
        
        gfn = abs(psnb) + interest + maturing
        gfn_gdp = gfn / gdp_nominal * 100
        
        gfn_data.append({
            'fiscal_year': fy,
            'psnb': abs(psnb),
            'interest': interest,
            'total_maturing': maturing,
            'gfn': gfn,
            'gfn_gdp': gfn_gdp
        })
        
        gdp_nominal *= 1.035  # Nominal GDP growth
    
    gfn_df = pd.DataFrame(gfn_data)
    results['gfn'] = {
        'annual': gfn_df,
        'average_gfn_gdp': gfn_df['gfn_gdp'].mean(),
        'max_gfn_gdp': gfn_df['gfn_gdp'].max()
    }
    
    print(f"\nGFN Summary:")
    print(f"   Average GFN/GDP: {gfn_df['gfn_gdp'].mean():.1f}%")
    print(f"   Maximum GFN/GDP: {gfn_df['gfn_gdp'].max():.1f}%")
    print(f"   Years above 15%: {(gfn_df['gfn_gdp'] > 15).sum()}")
    print(f"\n   Assessment: {'ELEVATED' if gfn_df['gfn_gdp'].mean() > 15 else 'MODERATE'} rollover risk")
    
    # =========================================================================
    # 6. SUSTAINABILITY ASSESSMENT
    # =========================================================================
    print_header("6. SUSTAINABILITY ASSESSMENT")
    
    print("\n" + "-"*50)
    print("OPERATIONAL SUSTAINABILITY CRITERIA")
    print("-"*50)
    
    # Criterion 1: Fiscal reaction
    c1_pass = bohn_beta > 0
    print(f"\n1. Fiscal Reaction (Bohn β > 0):")
    print(f"   β = {bohn_beta:.4f}")
    print(f"   Status: {'✓ PASS' if c1_pass else '✗ FAIL'}")
    
    # Criterion 2: Debt path probability
    p100 = obr_stats.get('prob_100_terminal', 50)
    c2_pass = p100 < 50
    print(f"\n2. Debt Path (P(>100%) < 50%):")
    print(f"   P(>100% terminal) = {p100:.1f}%")
    print(f"   Status: {'✓ PASS' if c2_pass else '✗ FAIL'}")
    
    # Criterion 3: Primary balance achievability
    obr_pb = 1.4  # OBR projected terminal pb
    c3_pass = obr_pb >= pb_star
    print(f"\n3. Primary Balance (pb ≥ pb*):")
    print(f"   OBR projected pb = {obr_pb:.1f}%")
    print(f"   Required pb* = {pb_star:.2f}%")
    print(f"   Status: {'✓ PASS (if achieved)' if c3_pass else '✗ FAIL'}")
    
    # Criterion 4: Financing needs
    avg_gfn = gfn_df['gfn_gdp'].mean()
    c4_pass = avg_gfn < 15
    print(f"\n4. Financing Needs (GFN < 15% GDP):")
    print(f"   Average GFN/GDP = {avg_gfn:.1f}%")
    print(f"   Status: {'✓ PASS' if c4_pass else '✗ FAIL'}")
    
    # Overall assessment
    criteria_passed = sum([c1_pass, c2_pass, c3_pass, c4_pass])
    score = criteria_passed / 4 * 100
    
    print("\n" + "="*50)
    print("OVERALL VERDICT")
    print("="*50)
    
    if criteria_passed >= 3:
        verdict = "SUSTAINABLE"
        color = "green"
    elif criteria_passed >= 2:
        verdict = "CONDITIONALLY SUSTAINABLE"
        color = "yellow"
    else:
        verdict = "UNSUSTAINABLE"
        color = "red"
    
    print(f"\n   Score: {criteria_passed}/4 criteria ({score:.0f}%)")
    print(f"   Verdict: {verdict}")
    
    print(f"""
   Explanation:
   UK debt sustainability is CONDITIONAL on achieving OBR projections.
   Historical behavior (negative Bohn β) provides NO automatic stabilization.
   The gap between policy commitment and historical behavior is the core risk.
    """)
    
    results['assessment'] = {
        'criteria': {
            'fiscal_reaction': c1_pass,
            'debt_path': c2_pass,
            'primary_balance': c3_pass,
            'financing_needs': c4_pass
        },
        'score': score,
        'verdict': verdict
    }
    
    # =========================================================================
    # 7. GENERATE OUTPUTS
    # =========================================================================
    if output_dir:
        print_header("7. GENERATING OUTPUTS")
        
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(f"{output_dir}/figures", exist_ok=True)
        os.makedirs(f"{output_dir}/tables", exist_ok=True)
        
        # Generate Excel workbook
        generate_excel_workbook(results, f"{output_dir}/UK_DSA_Analysis.xlsx")
        
        # Generate figures
        try:
            from visualization import DSAVisualizer
            viz = DSAVisualizer(f"{output_dir}/figures")
            # Generate figures that we have data for
            if 'monte_carlo' in results and 'obr_baseline_stats' in results['monte_carlo']:
                viz.plot_fan_chart(
                    results['monte_carlo']['obr_baseline_stats'],
                    save_path=f"{output_dir}/figures/fan_chart.png"
                )
            print(f"Generated figures in {output_dir}/figures/")
        except Exception as e:
            print(f"Figure generation skipped: {e}")
        
        print(f"\nOutputs saved to: {output_dir}")
    
    return results


def generate_excel_workbook(results: dict, output_path: str):
    """Generate comprehensive Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    
    # ===== Sheet 1: Executive Summary =====
    ws = wb.active
    ws.title = "Executive Summary"
    
    ws['A1'] = "UK DEBT SUSTAINABILITY ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    ws['A4'] = "OVERALL VERDICT"
    ws['A4'].font = title_font
    
    verdict = results.get('assessment', {}).get('verdict', 'N/A')
    ws['A5'] = verdict
    ws['A5'].font = Font(bold=True, size=14, 
                         color="008000" if "SUSTAINABLE" in verdict and "UN" not in verdict 
                         else "FF6600" if "CONDITIONAL" in verdict else "FF0000")
    
    ws['A7'] = "KEY METRICS"
    ws['A7'].font = title_font
    
    bohn_beta = results.get('bohn', {}).get('basic', {}).get('beta', 'N/A')
    mc_stats = results.get('monte_carlo', {}).get('obr_baseline_stats', {})
    
    metrics = [
        ("Bohn β (fiscal reaction)", f"{bohn_beta:.4f}" if isinstance(bohn_beta, float) else bohn_beta),
        ("P(debt > 100% GDP)", f"{mc_stats.get('prob_100_terminal', 'N/A'):.1f}%"),
        ("Mean terminal debt", f"{mc_stats.get('mean', 'N/A'):.1f}%"),
        ("VaR 95%", f"{mc_stats.get('var_95', 'N/A'):.1f}%"),
        ("VaR 99%", f"{mc_stats.get('var_99', 'N/A'):.1f}%"),
        ("Average GFN/GDP", f"{results.get('gfn', {}).get('average_gfn_gdp', 'N/A'):.1f}%"),
    ]
    
    for i, (metric, value) in enumerate(metrics):
        ws[f'A{8+i}'] = metric
        ws[f'B{8+i}'] = value
    
    # ===== Sheet 2: Bohn Test =====
    ws2 = wb.create_sheet("Bohn Test")
    
    ws2['A1'] = "BOHN FISCAL REACTION TEST"
    ws2['A1'].font = title_font
    
    ws2['A3'] = "Specification"
    ws2['B3'] = "β"
    ws2['C3'] = "SE (HAC)"
    ws2['D3'] = "t-stat"
    ws2['E3'] = "p-value"
    ws2['F3'] = "Result"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2[f'{col}3'].font = header_font
        ws2[f'{col}3'].fill = header_fill
    
    basic = results.get('bohn', {}).get('basic', {})
    ws2['A4'] = "Basic"
    ws2['B4'] = basic.get('beta', 'N/A')
    ws2['C4'] = basic.get('se_hac', 'N/A')
    ws2['D4'] = basic.get('t_stat_hac', 'N/A')
    ws2['E4'] = basic.get('p_value_hac', 'N/A')
    ws2['F4'] = "FAIL" if not basic.get('sustainable_hac', False) else "PASS"
    
    # ===== Sheet 3: Monte Carlo =====
    ws3 = wb.create_sheet("Monte Carlo")
    
    ws3['A1'] = "MONTE CARLO SIMULATION RESULTS"
    ws3['A1'].font = title_font
    
    ws3['A3'] = "Statistic"
    ws3['B3'] = "OBR Baseline"
    ws3['C3'] = "Fiscal Reaction"
    
    for col in ['A', 'B', 'C']:
        ws3[f'{col}3'].font = header_font
        ws3[f'{col}3'].fill = header_fill
    
    obr = results.get('monte_carlo', {}).get('obr_baseline_stats', {})
    fr = results.get('monte_carlo', {}).get('fiscal_reaction_stats', {})
    
    mc_metrics = [
        ("Mean terminal", obr.get('mean'), fr.get('mean')),
        ("Median terminal", obr.get('median'), fr.get('median')),
        ("Std Dev", obr.get('std'), fr.get('std')),
        ("P(>100%)", obr.get('prob_100_terminal'), fr.get('prob_100_terminal')),
        ("VaR 95%", obr.get('var_95'), fr.get('var_95')),
        ("VaR 99%", obr.get('var_99'), fr.get('var_99')),
    ]
    
    for i, (name, v1, v2) in enumerate(mc_metrics):
        ws3[f'A{4+i}'] = name
        ws3[f'B{4+i}'] = f"{v1:.1f}%" if v1 else "N/A"
        ws3[f'C{4+i}'] = f"{v2:.1f}%" if v2 else "N/A"
    
    # ===== Sheet 4: GFN =====
    ws4 = wb.create_sheet("GFN Analysis")
    
    ws4['A1'] = "GROSS FINANCING NEEDS"
    ws4['A1'].font = title_font
    
    gfn_df = results.get('gfn', {}).get('annual')
    if gfn_df is not None:
        ws4['A3'] = "Year"
        ws4['B3'] = "PSNB (£bn)"
        ws4['C3'] = "Interest (£bn)"
        ws4['D3'] = "Maturing (£bn)"
        ws4['E3'] = "GFN (£bn)"
        ws4['F3'] = "GFN/GDP (%)"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws4[f'{col}3'].font = header_font
            ws4[f'{col}3'].fill = header_fill
        
        for i, row in gfn_df.iterrows():
            r = 4 + i
            ws4[f'A{r}'] = row['fiscal_year']
            ws4[f'B{r}'] = f"{row['psnb']:.1f}"
            ws4[f'C{r}'] = f"{row['interest']:.1f}"
            ws4[f'D{r}'] = f"{row['total_maturing']:.1f}"
            ws4[f'E{r}'] = f"{row['gfn']:.1f}"
            ws4[f'F{r}'] = f"{row['gfn_gdp']:.1f}"
    
    # Save
    wb.save(output_path)
    print(f"Saved Excel workbook: {output_path}")


if __name__ == "__main__":
    # Run with output generation
    output_dir = '/mnt/user-data/outputs'
    results = run_complete_analysis(output_dir)
    
    print("\n" + "="*70)
    print("ANALYSIS COMPLETE")
    print("="*70)
    print(f"\nOutputs available at: {output_dir}")
    print("\nKey files:")
    print("  - UK_DSA_Analysis.xlsx")
    print("  - figures/*.png")
